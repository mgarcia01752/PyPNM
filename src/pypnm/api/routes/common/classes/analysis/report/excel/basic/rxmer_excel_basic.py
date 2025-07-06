# SPDX-License-Identifier: MIT
# Copyright (c) 2025 Maurice Garcia

import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Union
import json

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import NumericAxis
from openpyxl.utils import get_column_letter

from pypnm.api.routes.common.classes.analysis.analysis import Analysis
from pypnm.lib.mac_address import MacAddress


# Constants for chart positioning
CHART_START_COLUMN_INDEX = 4  # 'D'
CHART_START_ROW = 2

class RxMerExcelBasic:
    """
    Generates an XLSX RxMER Basic (PNN4) report, one sheet per analysis instance.
    Each sheet is named by its channel_id (e.g., "160").

    This version omits header rows on each sheet and writes only:
      - Two columns of data (frequency, magnitude)
      - A line chart with numeric X (frequency) and Y (magnitude) axes

    Attributes:
        analysis: Analysis object or list of dicts representing analysis instances.
        output_dir: Path where the XLSX file will be saved.
    """
    
    def __init__(self,
        analysis: Union[Analysis, List[Dict]],
        output_dir: Path):
        
        self.logger = logging.getLogger("RxMerExcelBasic")
        self.analysis = analysis
        self.output_dir = output_dir
        self.mac_address:MacAddress = None

    def build(self, filename: Optional[str] = None) -> Tuple[Path, List[str]]:
        """
        Build and save the XLSX report.

        Args:
            filename: Optional filename; if omitted, generates a timestamped name.

        Returns:
            Tuple of (output_path, sheet_names) for created sheets.
        """
        self.filename = filename
        workbook = Workbook()
        workbook.remove(workbook.active)

        sheet_names: List[str] = []

        # Get the raw list of instances
        if hasattr(self.analysis, "get_results"):
            result_dict = self.analysis.get_results()
            raw_instances = result_dict.get("analysis", []) if isinstance(result_dict, dict) else []
        else:
            raw_instances = self.analysis

        for item in raw_instances:
            inst = None

            if isinstance(item, str):

                try:
                    inst = json.loads(item)
                except json.JSONDecodeError:
                    self.logger.error("Could not parse analysis JSON", exc_info=True)
                    continue
                
            elif isinstance(item, dict):
                inst = item
                
            else:
                self.logger.error(f"Unexpected item type: {type(item)}")
                continue
            
            if not self.mac_address:
                self.mac_address = MacAddress(inst.get("mac_address", "00:00:00:00:00:00"))
                self.capture_time = inst.get("pnm_header",{}).get("capture_time", "0")
                            
            # Create sheet
            channel_id = inst.get("channel_id", "unknown")
            name = str(channel_id)
            sheet_names.append(name)
            ws = workbook.create_sheet(title=name)

            # Write data rows only: frequency (col 1), magnitude (col 2)
            freqs = inst.get("carrier_values", {}).get("frequency", [])
            mags = inst.get("carrier_values", {}).get("magnitude", [])
            row_count = min(len(freqs), len(mags))
            for i in range(row_count):
                ws.append([freqs[i], mags[i]])

            # Build line chart with numeric X-axis
            chart = LineChart()
            chart.title = f"RxMER Channel {channel_id}"

            # Data series: magnitude values in column 2
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=row_count)
            chart.add_data(data_ref, titles_from_data=False)

            # Categories: frequency values in column 1
            cat_ref = Reference(ws, min_col=1, min_row=1, max_row=row_count)
            chart.set_categories(cat_ref)

            # Numeric X-axis scaling
            chart.x_axis = NumericAxis()
            if row_count:
                f_min = min(freqs[:row_count])
                f_max = max(freqs[:row_count])
                chart.x_axis.scaling.min = f_min
                chart.x_axis.scaling.max = f_max
                chart.x_axis.majorUnit = (f_max - f_min) / 10
            chart.x_axis.title = inst.get("frequency_unit", "Frequency")
            chart.y_axis.title = inst.get("magnitude_unit", "Magnitude")

            # Position the chart
            col = get_column_letter(CHART_START_COLUMN_INDEX)
            ws.add_chart(chart, f"{col}{CHART_START_ROW}")

        # Ensure output directory
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        if not self.filename:
            dev_id = f'{self.mac_address.to_mac_format()}_{self.capture_time}'
            self.filename = f"rxmer_basic_{dev_id}.xlsx"
            
        out_path = self.output_dir / self.filename
        
        workbook.save(out_path)
        return out_path, sheet_names

    def get_filename(self) -> str:
        return self.filename