# SPDX-License-Identifier: MIT
# Copyright (c) 2025 Maurice Garcia

from openpyxl import Workbook
from openpyxl.styles import Font
from typing import List, Dict, Any, Optional
import os

class ExcelWorkbookFactory:
    """
    Base class to generate Excel workbooks with tables (sheets).
    Supports header formatting, autosizing, and workbook saving.
    """

    def __init__(self):
        self.workbook = Workbook()
        self._first_sheet_used = False

    def create_table(self,
        sheet_name: str,
        table_data: List[Dict[str, Any]],
        bold_headers: bool = True,
        auto_size: bool = True
    ):
        """
        Create a new sheet with tabular data.

        Args:
            sheet_name: Name of the Excel sheet (table).
            table_data: List of dictionaries, where each dict is a row.
            bold_headers: Whether to make headers bold.
            auto_size: Whether to autosize columns.
        """
        if self._first_sheet_used:
            sheet = self.workbook.create_sheet(title=sheet_name)
        else:
            sheet = self.workbook.active
            sheet.title = sheet_name
            self._first_sheet_used = True

        if not table_data:
            return

        headers = list(table_data[0].keys())

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            if bold_headers:
                cell.font = Font(bold=True)

        # Write rows
        for row_idx, row_data in enumerate(table_data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=row_data.get(header))

        # Auto-size columns
        if auto_size:
            for col_idx, header in enumerate(headers, start=1):
                max_length = max(len(str(header)), 12)
                for row in table_data:
                    val = str(row.get(header, ""))
                    max_length = max(max_length, len(val))
                sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = max_length + 2

    def save(self, path: str):
        """
        Save the workbook to the given path.

        Args:
            path: Full file path to save the workbook (should end in .xlsx)
        """
        os.makedirs(os.path.dirname(path), exist_ok=True)
        self.workbook.save(path)
